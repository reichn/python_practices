from fileinput import filename
from re import A
from openpyxl import Workbook
import datetime

t = datetime.date(2022, 7, 11)
f = "gantt.xlsx"

wb = Workbook()
ws = wb.active

ws["A1"] = "项目\日期"
ws["A2"] = "学习Javascript"
ws["B1"] = t

ws.freeze_panes = ws["B1"]

ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 15

wb.save(filename=f)
